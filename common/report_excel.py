"""
Excel workbook generation for report requests.
"""
from datetime import datetime
from pathlib import Path
from typing import Dict, List

from common.models import ReportDefinition, ReportRequestDetails, ResolvedReportDateRange


class ExcelReportGenerator:
    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(
        self,
        rows: List[Dict],
        definition: ReportDefinition,
        request: ReportRequestDetails,
        date_range: ResolvedReportDateRange,
    ) -> str:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        generated_at = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = self.output_dir / f"{definition.report_name}_{generated_at}.xlsx"
        dataframe = pd.DataFrame(rows)

        workbook = Workbook()
        data = workbook.active
        data.title = "Data"
        workbook.properties.title = definition.title
        workbook.properties.subject = f"{date_range.start_date} to {date_range.end_date}"
        workbook.properties.creator = "AMS AI"

        headers = list(dataframe.columns)
        if headers:
            header_fill = PatternFill("solid", fgColor="D9EAF7")
            for column_number, header in enumerate(headers, start=1):
                cell = data.cell(row=1, column=column_number, value=header)
                cell.font = Font(bold=True)
                cell.fill = header_fill
            for row_number, row in enumerate(dataframe.itertuples(index=False, name=None), start=2):
                for column_number, header in enumerate(headers, start=1):
                    data.cell(row=row_number, column=column_number, value=row[column_number - 1])
            data.freeze_panes = "A2"
            self._fit_columns(data, dataframe)
        else:
            data["A1"] = "No records returned"
            data["A1"].font = Font(italic=True)
            data.column_dimensions[get_column_letter(1)].width = 24

        workbook.save(output_path)
        return str(output_path)

    @staticmethod
    def _fit_columns(sheet, dataframe) -> None:
        from openpyxl.utils import get_column_letter

        for column_number, header in enumerate(dataframe.columns, start=1):
            values = [str(value) for value in dataframe[header].head(50).fillna("")]
            width = max([len(header)] + [len(value) for value in values])
            sheet.column_dimensions[get_column_letter(column_number)].width = min(max(width + 2, 12), 48)
