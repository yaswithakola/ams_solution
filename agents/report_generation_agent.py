"""
Report Generation AI Agent
==========================
Extracts structured report intent from a ServiceNow service request and
owns the deterministic report execution helpers used by the AMS flow.

This agent does not generate SQL and never receives database result data
for reasoning. It translates ticket text into validated parameters, then
runs only approved catalog-backed SQL through read-only access.
"""
import contextlib
import json
import logging
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import TYPE_CHECKING, Any, Dict, List, Optional

from common.models import (
    ReportDefinition,
    ReportExecutionResult,
    ReportFilterDefinition,
    ReportRequestDetails,
    ResolvedReportDateRange,
    Ticket,
)

if TYPE_CHECKING:
    from common.llm_client import AnthropicAgentClient
    from config import AnthropicSettings

logger = logging.getLogger(__name__)

KNOWN_REPORTS = {
    "claims_failure",
    "member_enrollment",
    "billing_summary",
    "batch_job_failure",
}
KNOWN_FREQUENCIES = {"daily", "weekly", "monthly", "adhoc"}
KNOWN_DATE_RANGES = {"today", "yesterday", "previous_week", "previous_month", "custom"}
_BLOCKED_SQL = re.compile(
    r"\b(ALTER|CALL|CREATE|DELETE|DROP|EXEC|EXECUTE|GRANT|INSERT|MERGE|REVOKE|TRUNCATE|UPDATE)\b",
    re.IGNORECASE,
)

SYSTEM_PROMPT = """You are the Report Generation AI Agent for an AMS service catalog flow. \
Extract the user's report request into structured fields only.

Approved reports:
- claims_failure
- member_enrollment
- billing_summary
- batch_job_failure

Rules:
- Do not generate SQL.
- Do not execute a report.
- Do not calculate actual dates. Use date_range labels only: today, yesterday, previous_week, \
  previous_month, or custom. Only return start_date/end_date when the user gives explicit custom dates.
- Put business filters such as state=Texas in the filters object.
- If the request does not name an email recipient, use the ServiceNow requester email when it is available.
- Use output_format="excel" unless the user explicitly asks for another format. If they do, still \
  return excel and explain the limitation in rationale.
- If the report name cannot be matched to the approved report list, set insufficient_information=true.

Return strict JSON with exactly these keys: report_name, frequency, date_range, start_date, end_date, \
filters, output_format, recipient, rationale, insufficient_information."""


class ReportGenerationAgent:
    def __init__(self, llm_client: "AnthropicAgentClient", anthropic_settings: "AnthropicSettings"):
        self.llm_client = llm_client
        self.model = anthropic_settings.model_report_request_parser

    def parse(self, ticket: Ticket) -> ReportRequestDetails:
        user_prompt = f"""Service request:
Catalog task number: {ticket.number}
Catalog task short description: {ticket.short_description}
Catalog task description: {ticket.description}
Parent RITM number: {ticket.request_item_number or "unknown"}
Parent RITM short description: {ticket.request_item_short_description or "unknown"}
Parent RITM description: {ticket.request_item_description or "unknown"}
Requested for email: {ticket.requested_for_email or "unknown"}
Opened by email: {ticket.opened_by_email or "unknown"}

Extract the report request fields."""

        result = self.llm_client.complete_json(SYSTEM_PROMPT, user_prompt, model=self.model, temperature=0.1)

        report_name = self._normalize_report_name(result.get("report_name"))
        frequency = self._normalize_choice(result.get("frequency"), KNOWN_FREQUENCIES, default="adhoc")
        date_range = self._normalize_choice(result.get("date_range"), KNOWN_DATE_RANGES, default=None)
        start_date = result.get("start_date") if isinstance(result.get("start_date"), str) else None
        end_date = result.get("end_date") if isinstance(result.get("end_date"), str) else None
        filters = self._clean_filters(result.get("filters"))
        output_format = self._normalize_choice(result.get("output_format"), {"excel"}, default="excel")
        recipient = result.get("recipient") if isinstance(result.get("recipient"), str) else None
        if not recipient:
            recipient = ticket.requested_for_email or ticket.opened_by_email

        insufficient = bool(result.get("insufficient_information", False))
        if report_name not in KNOWN_REPORTS or date_range is None:
            insufficient = True

        details = ReportRequestDetails(
            ticket_number=ticket.number,
            report_name=report_name if report_name in KNOWN_REPORTS else None,
            frequency=frequency,
            date_range=date_range,
            start_date=start_date,
            end_date=end_date,
            filters=filters,
            output_format=output_format,
            recipient=recipient.strip() if recipient else None,
            rationale=result.get("rationale", ""),
            insufficient_information=insufficient,
        )

        logger.info(
            "Parsed report request %s: report=%s frequency=%s date_range=%s insufficient=%s",
            ticket.number,
            details.report_name,
            details.frequency,
            details.date_range,
            details.insufficient_information,
        )
        return details

    @staticmethod
    def _normalize_report_name(value) -> str:
        if not isinstance(value, str):
            return ""
        return normalize_report_name(value)

    @staticmethod
    def _normalize_choice(value, allowed, default=None):
        if not isinstance(value, str):
            return default
        normalized = normalize_report_name(value)
        aliases = {
            "last_week": "previous_week",
            "last_month": "previous_month",
            "on_demand": "adhoc",
            "ad_hoc": "adhoc",
        }
        normalized = aliases.get(normalized, normalized)
        return normalized if normalized in allowed else default

    @staticmethod
    def _clean_filters(value) -> dict:
        if not isinstance(value, dict):
            return {}
        filters = {}
        for key, item in value.items():
            if not isinstance(key, str) or not key.strip():
                continue
            if item is None:
                continue
            if isinstance(item, (str, int, float, bool)):
                filters[key.strip().lower()] = str(item).strip()
        return filters


class ReportCatalog:
    def __init__(self, catalog_path: str, sql_dir: str):
        self.catalog_path = Path(catalog_path)
        self.sql_dir = Path(sql_dir)
        self._reports = self._load_reports()

    def get(self, report_name: str) -> ReportDefinition:
        key = normalize_report_name(report_name)
        if key not in self._reports:
            raise KeyError(f"Unknown report: {report_name}")
        return self._reports[key]

    def load_sql(self, definition: ReportDefinition) -> str:
        sql_path = (self.sql_dir / definition.sql_file).resolve()
        sql_root = self.sql_dir.resolve()

        if sql_path != sql_root and sql_root not in sql_path.parents:
            raise ValueError(f"SQL file escapes report SQL directory: {definition.sql_file}")

        sql = sql_path.read_text(encoding="utf-8")
        validate_select_only_sql(sql)
        return sql

    def _load_reports(self) -> Dict[str, ReportDefinition]:
        raw = json.loads(self.catalog_path.read_text(encoding="utf-8"))
        reports = {}
        for item in raw.get("reports", []):
            definition = ReportDefinition(**item)
            reports[normalize_report_name(definition.report_name)] = definition
        return reports


class ReportDateResolver:
    def __init__(self, today: Optional[date] = None):
        self.today = today or date.today()

    def resolve(self, request: ReportRequestDetails) -> ResolvedReportDateRange:
        label = normalize_report_name(request.date_range or "")
        aliases = {
            "last_week": "previous_week",
            "last_month": "previous_month",
        }
        label = aliases.get(label, label)

        if label == "today":
            return ResolvedReportDateRange(start_date=self.today, end_date=self.today, label=label)

        if label == "yesterday":
            yesterday = self.today - timedelta(days=1)
            return ResolvedReportDateRange(start_date=yesterday, end_date=yesterday, label=label)

        if label == "previous_week":
            this_week_start = self.today - timedelta(days=self.today.weekday())
            start = this_week_start - timedelta(days=7)
            end = this_week_start - timedelta(days=1)
            return ResolvedReportDateRange(start_date=start, end_date=end, label=label)

        if label == "previous_month":
            first_this_month = self.today.replace(day=1)
            end = first_this_month - timedelta(days=1)
            start = end.replace(day=1)
            return ResolvedReportDateRange(start_date=start, end_date=end, label=label)

        if label == "custom":
            if not request.start_date or not request.end_date:
                raise ValueError("Custom report date range requires start_date and end_date")
            start = date.fromisoformat(request.start_date)
            end = date.fromisoformat(request.end_date)
            if start > end:
                raise ValueError("Report start_date must be on or before end_date")
            return ResolvedReportDateRange(start_date=start, end_date=end, label=label)

        raise ValueError(f"Unsupported report date range: {request.date_range}")


class PostgresReportClient:
    def __init__(self, dsn: str):
        self.dsn = dsn

    def query(self, sql: str, params: Dict[str, Any]) -> List[Dict[str, Any]]:
        validate_select_only_sql(sql)
        with contextlib.closing(self._connect()) as conn:
            conn.set_session(readonly=True, autocommit=False)
            with conn.cursor() as cur:
                cur.execute(sql, params)
                rows = cur.fetchall()
            conn.rollback()
        return [dict(row) for row in rows]

    def _connect(self):
        import psycopg2
        import psycopg2.extras

        return psycopg2.connect(self.dsn, cursor_factory=psycopg2.extras.RealDictCursor)


class ExcelReportGenerator:
    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(
        self,
        rows: List[Dict[str, Any]],
        definition: ReportDefinition,
        request: ReportRequestDetails,
        date_range: ResolvedReportDateRange,
    ) -> str:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        generated_at = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = self.output_dir / f"{definition.report_name}_{generated_at}.xlsx"
        dataframe = pd.DataFrame(rows)

        workbook = Workbook()
        data = workbook.active
        data.title = "Data"
        workbook.properties.title = definition.title
        workbook.properties.subject = f"{date_range.start_date} to {date_range.end_date}"
        workbook.properties.creator = "AMS AI"

        headers = list(dataframe.columns)
        if headers:
            header_fill = PatternFill("solid", fgColor="D9EAF7")
            for column_number, header in enumerate(headers, start=1):
                cell = data.cell(row=1, column=column_number, value=header)
                cell.font = Font(bold=True)
                cell.fill = header_fill
            for row_number, row in enumerate(dataframe.itertuples(index=False, name=None), start=2):
                for column_number, _header in enumerate(headers, start=1):
                    data.cell(row=row_number, column=column_number, value=row[column_number - 1])
            data.freeze_panes = "A2"
            self._fit_columns(data, dataframe)
        else:
            data["A1"] = "No records returned"
            data["A1"].font = Font(italic=True)
            data.column_dimensions[get_column_letter(1)].width = 24

        workbook.save(output_path)
        return str(output_path)

    @staticmethod
    def _fit_columns(sheet, dataframe) -> None:
        from openpyxl.utils import get_column_letter

        for column_number, header in enumerate(dataframe.columns, start=1):
            values = [str(value) for value in dataframe[header].head(50).fillna("")]
            width = max([len(header)] + [len(value) for value in values])
            sheet.column_dimensions[get_column_letter(column_number)].width = min(max(width + 2, 12), 48)


class ReportService:
    def __init__(
        self,
        catalog: ReportCatalog,
        database_client,
        date_resolver: Optional[ReportDateResolver] = None,
        excel_generator: Optional[ExcelReportGenerator] = None,
        ses_settings=None,
        email_sender=None,
    ):
        self.catalog = catalog
        self.database_client = database_client
        self.date_resolver = date_resolver or ReportDateResolver()
        self.excel_generator = excel_generator or ExcelReportGenerator()
        self.ses_settings = ses_settings
        self.email_sender = email_sender

    def run(self, request: ReportRequestDetails) -> ReportExecutionResult:
        try:
            definition = self.catalog.get(request.report_name)
            date_range = self.date_resolver.resolve(request)
            params = build_report_parameters(request, definition, date_range)
            sql = self.catalog.load_sql(definition)

            logger.info(
                "Running report %s for %s to %s",
                definition.report_name,
                date_range.start_date,
                date_range.end_date,
            )
            rows = self.database_client.query(sql, params)
            recipient = request.recipient or definition.default_recipient
            output_path = self.excel_generator.generate(rows, definition, request, date_range)
            email_sent = False
            message = "Excel report generated."

            if recipient and (self.ses_settings or self.email_sender):
                email_sender = self.email_sender
                if email_sender is None:
                    from common.email_utils import send_report_email_ses

                    email_sender = send_report_email_ses

                email_sent = email_sender(
                    self.ses_settings,
                    recipient=recipient,
                    subject=f"{request.ticket_number} - {definition.title}",
                    text_body=(
                        f"Hello,\n\n"
                        f"The {definition.title} for {date_range.label.replace('_', ' ')} has been generated.\n"
                        f"Please find the report attached.\n\n"
                        f"Request number: {request.ticket_number}\n\n"
                        f"Regards,\n"
                        f"{_display_sender_name(self.ses_settings)}"
                    ),
                    attachment_path=output_path,
                )
                message = "Excel report generated and email sent through SES." if email_sent else (
                    "Excel report generated, but email was not sent. Check SES configuration."
                )

            return ReportExecutionResult(
                ticket_number=request.ticket_number,
                report_name=definition.report_name,
                status="EMAIL_SENT" if email_sent else "REPORT_GENERATED",
                record_count=len(rows),
                recipient=recipient,
                message=message,
                output_path=output_path,
                email_sent=email_sent,
            )
        except Exception as exc:
            logger.exception("Report request %s failed", request.ticket_number)
            return ReportExecutionResult(
                ticket_number=request.ticket_number,
                report_name=request.report_name or "unknown",
                status="FAILED",
                message=str(exc),
            )


def normalize_report_name(value: str) -> str:
    return (value or "").strip().lower().replace("-", "_").replace(" ", "_")


def _display_sender_name(ses_settings) -> str:
    source_email = getattr(ses_settings, "source_email", "") if ses_settings else ""
    if not source_email or "@" not in source_email:
        return "Support Team"
    local_part = source_email.split("@", 1)[0]
    cleaned = re.sub(r"[0-9._-]+", " ", local_part).strip()
    if not cleaned:
        return "Support Team"
    return " ".join(part.capitalize() for part in cleaned.split())


def validate_select_only_sql(sql: str) -> str:
    cleaned = _strip_sql_comments(sql).strip()
    if not cleaned:
        raise ValueError("Report SQL is empty")

    if ";" in cleaned.rstrip(";"):
        raise ValueError("Report SQL must contain a single statement")

    statement = cleaned.rstrip(";").lstrip()
    if not re.match(r"^(SELECT|WITH)\b", statement, flags=re.IGNORECASE):
        raise ValueError("Report SQL must start with SELECT or WITH")

    if _BLOCKED_SQL.search(statement):
        raise ValueError("Report SQL contains a blocked write/admin statement")

    return sql


def build_report_parameters(
    request: ReportRequestDetails,
    definition: ReportDefinition,
    date_range: ResolvedReportDateRange,
) -> Dict[str, Any]:
    allowed_filters = {item.name: item for item in definition.allowed_filters}
    unknown_filters = sorted(set(request.filters) - set(allowed_filters))
    if unknown_filters:
        raise ValueError(f"Unsupported filters for {definition.report_name}: {', '.join(unknown_filters)}")

    params = {
        "start_date": date_range.start_date,
        "end_date": date_range.end_date,
    }
    for filter_definition in definition.allowed_filters:
        raw_value = request.filters.get(filter_definition.name)
        if raw_value in (None, "") and filter_definition.required:
            raise ValueError(f"Missing required filter: {filter_definition.name}")
        params[filter_definition.name] = _coerce_filter(raw_value, filter_definition)
    return params


def _coerce_filter(value: Any, filter_definition: ReportFilterDefinition) -> Any:
    if value in (None, ""):
        return None

    if filter_definition.value_type == "string":
        result = str(value).strip()
    elif filter_definition.value_type == "integer":
        result = int(value)
    elif filter_definition.value_type == "number":
        result = float(value)
    elif filter_definition.value_type == "boolean":
        if isinstance(value, bool):
            result = value
        elif str(value).strip().lower() in ("true", "false"):
            result = str(value).strip().lower() == "true"
        else:
            raise ValueError(f"Filter {filter_definition.name} must be true or false")
    else:
        raise ValueError(f"Unsupported filter type: {filter_definition.value_type}")

    if filter_definition.allowed_values and result not in filter_definition.allowed_values:
        raise ValueError(f"Filter {filter_definition.name} has unsupported value: {result}")
    return result


def _strip_sql_comments(sql: str) -> str:
    no_block_comments = re.sub(r"/\*.*?\*/", "", sql, flags=re.DOTALL)
    return re.sub(r"--.*?$", "", no_block_comments, flags=re.MULTILINE)
