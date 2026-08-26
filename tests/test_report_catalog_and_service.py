"""
Offline tests for approved report lookup, date handling, SQL parameters,
and report service execution.
"""
from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import Mock

from agents.report_generation_agent import (
    ExcelReportGenerator,
    ReportCatalog,
    ReportDateResolver,
    ReportService,
    build_report_parameters,
    validate_select_only_sql,
)
from common.models import ReportRequestDetails


class FakeReportDatabase:
    def __init__(self):
        self.sql = None
        self.params = None

    def query(self, sql, params):
        self.sql = sql
        self.params = params
        return [
            {
                "member_id": "M1001",
                "state": "Texas",
                "enrollment_status": "ACTIVE",
            }
        ]


class FakeSesSettings:
    region_name = "us-east-1"
    source_email = "reports@example.com"
    configuration_set = ""


def make_catalog():
    return ReportCatalog("reporting/reports.json", "reporting/sql")


def test_report_catalog_lookup_and_sql_load():
    catalog = make_catalog()
    definition = catalog.get("member enrollment")
    sql = catalog.load_sql(definition)

    assert definition.report_name == "member_enrollment"
    assert "FROM member_enrollment" in sql
    print("test_report_catalog_lookup_and_sql_load: PASSED")


def test_date_resolver_previous_month():
    request = ReportRequestDetails(
        ticket_number="RITM001",
        report_name="member_enrollment",
        frequency="monthly",
        date_range="previous_month",
    )
    resolved = ReportDateResolver(today=date(2026, 8, 25)).resolve(request)

    assert resolved.start_date == date(2026, 7, 1)
    assert resolved.end_date == date(2026, 7, 31)
    print("test_date_resolver_previous_month: PASSED")


def test_date_resolver_previous_week():
    request = ReportRequestDetails(
        ticket_number="RITM002",
        report_name="claims_failure",
        frequency="weekly",
        date_range="previous_week",
    )
    resolved = ReportDateResolver(today=date(2026, 8, 25)).resolve(request)

    assert resolved.start_date == date(2026, 8, 17)
    assert resolved.end_date == date(2026, 8, 23)
    print("test_date_resolver_previous_week: PASSED")


def test_sql_validation_blocks_write_statements():
    try:
        validate_select_only_sql("DELETE FROM member_enrollment")
    except ValueError as exc:
        assert "SELECT" in str(exc) or "blocked" in str(exc)
        print("test_sql_validation_blocks_write_statements: PASSED")
        return
    raise AssertionError("DELETE statement was not rejected")


def test_report_parameters_allow_only_catalog_filters():
    catalog = make_catalog()
    definition = catalog.get("member_enrollment")
    request = ReportRequestDetails(
        ticket_number="RITM003",
        report_name="member_enrollment",
        frequency="monthly",
        date_range="previous_month",
        filters={"state": "Texas"},
    )
    resolved = ReportDateResolver(today=date(2026, 8, 25)).resolve(request)
    params = build_report_parameters(request, definition, resolved)

    assert params["start_date"] == date(2026, 7, 1)
    assert params["end_date"] == date(2026, 7, 31)
    assert params["state"] == "Texas"

    bad = request.model_copy(update={"filters": {"sql": "DROP TABLE x"}})
    try:
        build_report_parameters(bad, definition, resolved)
    except ValueError as exc:
        assert "Unsupported filters" in str(exc)
        print("test_report_parameters_allow_only_catalog_filters: PASSED")
        return
    raise AssertionError("Unsupported filter was not rejected")


def test_report_service_runs_approved_query():
    database = FakeReportDatabase()
    with TemporaryDirectory() as tmpdir:
        service = ReportService(
            catalog=make_catalog(),
            database_client=database,
            date_resolver=ReportDateResolver(today=date(2026, 8, 25)),
            excel_generator=ExcelReportGenerator(tmpdir),
        )
        request = ReportRequestDetails(
            ticket_number="RITM004",
            report_name="member_enrollment",
            frequency="monthly",
            date_range="previous_month",
            filters={"state": "Texas"},
            recipient="ops@example.com",
        )

        result = service.run(request)

        assert result.status == "REPORT_GENERATED"
        assert result.record_count == 1
        assert result.recipient == "ops@example.com"
        assert Path(result.output_path).exists()
        assert database.params["state"] == "Texas"
        assert "FROM member_enrollment" in database.sql
    print("test_report_service_runs_approved_query: PASSED")


def test_excel_workbook_contains_data_only():
    from openpyxl import load_workbook

    database = FakeReportDatabase()
    with TemporaryDirectory() as tmpdir:
        service = ReportService(
            catalog=make_catalog(),
            database_client=database,
            date_resolver=ReportDateResolver(today=date(2026, 8, 25)),
            excel_generator=ExcelReportGenerator(tmpdir),
        )
        request = ReportRequestDetails(
            ticket_number="RITM005",
            report_name="member_enrollment",
            frequency="monthly",
            date_range="previous_month",
            filters={"state": "Texas"},
        )
        result = service.run(request)

        workbook = load_workbook(result.output_path)
        assert workbook["Data"]["A1"].value == "member_id"
        assert workbook["Data"]["A2"].value == "M1001"
        assert workbook.sheetnames == ["Data"]
    print("test_excel_workbook_contains_data_only: PASSED")


def test_report_service_sends_email_through_ses_when_configured():
    database = FakeReportDatabase()
    with TemporaryDirectory() as tmpdir:
        mock_send = Mock(return_value=True)
        service = ReportService(
            catalog=make_catalog(),
            database_client=database,
            date_resolver=ReportDateResolver(today=date(2026, 8, 25)),
            excel_generator=ExcelReportGenerator(tmpdir),
            ses_settings=FakeSesSettings(),
            email_sender=mock_send,
        )
        request = ReportRequestDetails(
            ticket_number="RITM007",
            report_name="member_enrollment",
            frequency="monthly",
            date_range="previous_month",
            filters={"state": "Texas"},
            recipient="ops@example.com",
        )

        result = service.run(request)

        assert result.status == "EMAIL_SENT"
        assert result.email_sent is True
        mock_send.assert_called_once()
        args, kwargs = mock_send.call_args
        assert args[0] is not None
        assert kwargs["recipient"] == "ops@example.com"
        assert kwargs["attachment_path"] == result.output_path
    print("test_report_service_sends_email_through_ses_when_configured: PASSED")


if __name__ == "__main__":
    test_report_catalog_lookup_and_sql_load()
    test_date_resolver_previous_month()
    test_date_resolver_previous_week()
    test_sql_validation_blocks_write_statements()
    test_report_parameters_allow_only_catalog_filters()
    test_report_service_runs_approved_query()
    test_excel_workbook_contains_data_only()
    test_report_service_sends_email_through_ses_when_configured()
    print("All report catalog/service tests passed.")
